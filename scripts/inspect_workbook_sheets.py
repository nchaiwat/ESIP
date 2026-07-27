from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    for path in sorted((root / "SourceFiles").glob("*/incoming/*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        print(f"\n=== {path.relative_to(root)} ===")
        for sheet in workbook:
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = list(next(rows))
            except StopIteration:
                headers = []
            print(sheet.title, sheet.max_row, sheet.max_column, headers[:20])
        workbook.close()


if __name__ == "__main__":
    main()
