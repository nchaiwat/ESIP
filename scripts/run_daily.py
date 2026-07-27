from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class StepResult:
    name: str
    status: str
    return_code: int
    output: str
    elapsed_seconds: float = 0.0


EXPECTED_PREVIEW_SHEETS = {
    "Dashboard",
    "Daily Action List",
    "Coverage",
    "Input Freshness",
    "Input File Safety",
    "Manual Report Coverage",
    "Batch History",
    "Daily Trend",
    "Rankings",
    "Quarantine",
    "Product Mapping",
    "Branch Approval",
    "Publication Readiness",
    "Mapping Action Plan",
    "Approval Instructions",
}


def _prune_old_files(directory: Path, pattern: str, keep: int) -> int:
    if keep < 1 or not directory.is_dir():
        return 0
    files = sorted(
        directory.glob(pattern),
        key=lambda path: (path.stat().st_mtime_ns, path.name),
        reverse=True,
    )
    for path in files[keep:]:
        path.unlink()
    return max(0, len(files) - keep)


def _run(root: Path, name: str, command: list[str]) -> StepResult:
    started = time.perf_counter()
    completed = subprocess.run(
        command,
        cwd=root,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
    )
    elapsed = time.perf_counter() - started
    output = "\n".join(part.strip() for part in (completed.stdout, completed.stderr) if part.strip())
    return StepResult(
        name,
        "PASS" if completed.returncode == 0 else "FAIL",
        completed.returncode,
        output,
        round(elapsed, 3),
    )


def _find_node_runtime() -> tuple[Path, Path] | None:
    configured_node = os.environ.get("ESIP_NODE_EXE")
    configured_modules = os.environ.get("ESIP_NODE_MODULES")
    if configured_node and configured_modules:
        node = Path(configured_node)
        modules = Path(configured_modules)
        if node.is_file() and modules.is_dir():
            return node, modules

    cache_root = Path.home() / ".cache" / "codex-runtimes"
    candidates = sorted(
        cache_root.glob("*/dependencies/node/bin/node.exe"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for node in candidates:
        modules = node.parents[1] / "node_modules"
        if modules.is_dir():
            return node, modules
    return None


def _preview_is_valid(path: Path) -> tuple[bool, str]:
    if not path.is_file() or path.stat().st_size == 0:
        return False, "Preview workbook was not created"
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        names = set(workbook.sheetnames)
        workbook.close()
    except Exception as error:
        return False, f"Preview workbook cannot be opened: {error}"
    missing = sorted(EXPECTED_PREVIEW_SHEETS - names)
    if missing:
        return False, f"Preview workbook is missing sheets: {', '.join(missing)}"
    return True, "Preview workbook structure verified"


def _write_report(
    root: Path,
    run_id: str,
    started_at: str,
    finished_at: str,
    steps: list[StepResult],
    preview_path: Path | None,
    archived_preview: Path | None,
) -> Path:
    report_dir = root / "output" / "daily_runs"
    report_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "run_id": run_id,
        "started_at": started_at,
        "finished_at": finished_at,
        "overall_status": "PASS" if all(step.status == "PASS" for step in steps) else "FAIL",
        "preview_path": str(preview_path.relative_to(root)) if preview_path else None,
        "archived_preview": (
            str(archived_preview.relative_to(root)) if archived_preview else None
        ),
        "steps": [asdict(step) for step in steps],
    }
    json_path = report_dir / f"daily_run_{run_id}.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        f"# ESIP Daily Run {run_id}",
        "",
        f"- Overall: **{payload['overall_status']}**",
        f"- Started: {started_at}",
        f"- Finished: {finished_at}",
        f"- Preview: `{payload['preview_path'] or 'not created'}`",
        f"- Archived previous preview: `{payload['archived_preview'] or 'none'}`",
        "",
        "## Steps",
        "",
    ]
    for step in steps:
        lines.append(f"- **{step.status}** - {step.name} ({step.elapsed_seconds:.3f}s)")
    markdown_path = report_dir / f"daily_run_{run_id}.md"
    markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    (report_dir / "latest_run.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (report_dir / "latest_run.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return markdown_path


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    python = Path(sys.executable)
    started = datetime.now().astimezone()
    run_id = started.strftime("%Y%m%d_%H%M%S")
    steps: list[StepResult] = []
    preview_path: Path | None = None
    archived_preview: Path | None = None

    commands = [
        ("Refresh governed input manifest", [str(python), "scripts/refresh_manifest.py"]),
        ("Verify governed input manifest", [str(python), "-m", "esip.cli", "verify-manifest"]),
        ("Check PostgreSQL", [str(python), "-m", "esip.cli", "postgres-status"]),
        (
            "Synchronize SAP reference masters",
            [str(python), "-m", "esip.cli", "postgres-load-master"],
        ),
        ("Check OSCN reprocess", [str(python), "scripts/reprocess_after_oscn.py"]),
        ("Load HP and MH Daily Raw", [str(python), "-m", "esip.cli", "postgres-load-hp-mh"]),
        ("Load TWD Daily Raw", [str(python), "-m", "esip.cli", "postgres-load-twd"]),
        ("Load HH Daily Raw", [str(python), "-m", "esip.cli", "postgres-load-hh"]),
        ("Load GBH Daily Raw", [str(python), "-m", "esip.cli", "postgres-load-gbh"]),
        ("Load DH Daily Raw", [str(python), "-m", "esip.cli", "postgres-load-dh"]),
        ("Export mapping work queues", [str(python), "scripts/export_mapping_work_queues.py"]),
        ("Validate mapping candidates", [str(python), "scripts/check_mapping_candidates.py"]),
        ("Export publication readiness", [str(python), "scripts/export_publication_queue.py"]),
        ("Export Preview data", [str(python), "scripts/export_preview_data.py"]),
    ]
    for name, command in commands:
        result = _run(root, name, command)
        steps.append(result)
        print(f"{result.status}: {name} ({result.elapsed_seconds:.3f}s)")
        if result.status == "FAIL":
            break

    if all(step.status == "PASS" for step in steps) and len(steps) == len(commands):
        runtime = _find_node_runtime()
        if runtime is None:
            steps.append(
                StepResult(
                    "Build Excel Preview",
                    "FAIL",
                    1,
                    "Bundled Node runtime was not found. Set ESIP_NODE_EXE and ESIP_NODE_MODULES.",
                )
            )
        else:
            node, modules = runtime
            junction = root / "node_modules"
            created_junction = False
            temp_dir = root / ".tmp_daily"
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_output = temp_dir / f"ESIP_Daily_Raw_Preview_{run_id}.xlsx"
            try:
                if not junction.exists():
                    link = _run(
                        root,
                        "Prepare spreadsheet runtime",
                        ["cmd.exe", "/c", "mklink", "/J", str(junction), str(modules)],
                    )
                    steps.append(link)
                    created_junction = link.status == "PASS"
                if junction.exists():
                    build = _run(
                        root,
                        "Build Excel Preview",
                        [
                            str(node),
                            "scripts/build_daily_raw_preview.mjs",
                            str(root / ".tmp_review" / "preview_data.json"),
                            str(temp_output),
                            str(temp_dir / "dashboard.png"),
                        ],
                    )
                    valid, detail = _preview_is_valid(temp_output)
                    if valid:
                        build = StepResult(
                            "Build Excel Preview",
                            "PASS",
                            0,
                            detail,
                            build.elapsed_seconds,
                        )
                    steps.append(build)
                    if valid:
                        stable = root / "output" / "reports" / "ESIP_Daily_Raw_Preview.xlsx"
                        if stable.is_file():
                            archive_dir = root / "output" / "reports" / "archive"
                            archive_dir.mkdir(parents=True, exist_ok=True)
                            archived_preview = archive_dir / (
                                f"ESIP_Daily_Raw_Preview_before_{run_id}.xlsx"
                            )
                            shutil.copy2(stable, archived_preview)
                        stable.parent.mkdir(parents=True, exist_ok=True)
                        temp_output.replace(stable)
                        preview_path = stable
            finally:
                if created_junction and junction.exists():
                    subprocess.run(
                        ["cmd.exe", "/c", "rmdir", str(junction)],
                        cwd=root,
                        capture_output=True,
                        check=False,
                    )
                shutil.rmtree(temp_dir, ignore_errors=True)
                shutil.rmtree(root / ".tmp_review", ignore_errors=True)

    finished = datetime.now().astimezone()
    report = _write_report(
        root,
        run_id,
        started.isoformat(timespec="seconds"),
        finished.isoformat(timespec="seconds"),
        steps,
        preview_path,
        archived_preview,
    )
    removed_archives = _prune_old_files(
        root / "output" / "reports" / "archive",
        "ESIP_Daily_Raw_Preview_before_*.xlsx",
        30,
    )
    report_dir = root / "output" / "daily_runs"
    removed_reports = _prune_old_files(report_dir, "daily_run_*.md", 90)
    removed_reports += _prune_old_files(report_dir, "daily_run_*.json", 90)
    if removed_archives or removed_reports:
        print(
            f"Retention cleanup: {removed_archives} Preview archive(s), "
            f"{removed_reports} Daily Run report file(s)"
        )
    overall_pass = all(step.status == "PASS" for step in steps) and preview_path is not None
    print(f"{'PASS' if overall_pass else 'FAIL'}: Daily Run")
    print(f"Report: {report.relative_to(root)}")
    if preview_path:
        print(f"Preview: {preview_path.relative_to(root)}")
    return 0 if overall_pass else 1


if __name__ == "__main__":
    raise SystemExit(main())
