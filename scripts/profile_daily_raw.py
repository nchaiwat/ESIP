from __future__ import annotations

import csv
import json
from pathlib import Path
from zipfile import ZipFile

import xlrd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def clean(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def profile_xlsx(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        declared_rows = sheet.max_row or 0
        sample = [
            [clean(value) for value in row]
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(declared_rows, 15) if declared_rows else 15,
                values_only=True,
            )
        ]
        sheets.append(
            {
                "name": sheet.title,
                "rows": declared_rows,
                "columns": sheet.max_column or max((len(row) for row in sample), default=0),
                "sample": sample,
            }
        )
    workbook.close()
    return {"format": "xlsx", "sheets": sheets}


def profile_xls(path: Path) -> dict[str, object]:
    workbook = xlrd.open_workbook(path, on_demand=True)
    sheets = []
    for name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(name)
        sample = [
            [clean(sheet.cell_value(row, column)) for column in range(sheet.ncols)]
            for row in range(min(sheet.nrows, 15))
        ]
        sheets.append(
            {"name": name, "rows": sheet.nrows, "columns": sheet.ncols, "sample": sample}
        )
    workbook.release_resources()
    return {"format": "xls", "sheets": sheets}


def profile_zip(path: Path) -> dict[str, object]:
    members = []
    with ZipFile(path) as archive:
        for name in archive.namelist():
            raw = archive.read(name)
            encoding = "utf-8-sig"
            try:
                text = raw.decode(encoding)
            except UnicodeDecodeError:
                encoding = "cp874"
                text = raw.decode(encoding)
            rows = list(csv.reader(text.splitlines()))[:15]
            members.append(
                {
                    "name": name,
                    "bytes": len(raw),
                    "encoding": encoding,
                    "sample": rows,
                }
            )
    return {"format": "zip_csv", "members": members}


def main() -> None:
    profiles: dict[str, object] = {}
    for path in sorted((ROOT / "SourceFiles").glob("*/incoming/*")):
        if not path.is_file() or path.name == ".gitkeep" or "Unclassified" in str(path):
            continue
        if path.suffix.lower() == ".xlsx":
            details = profile_xlsx(path)
        elif path.suffix.lower() == ".xls":
            details = profile_xls(path)
        elif path.suffix.lower() == ".zip":
            details = profile_zip(path)
        else:
            continue
        profiles[str(path.relative_to(ROOT))] = details
    output = ROOT / "output" / "raw_profiles" / "daily_raw_inventory.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(profiles, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
