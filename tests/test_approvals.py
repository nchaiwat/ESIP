from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from esip import approvals
from esip.approvals import ApprovalResult, apply_approval_result
from esip.master_data import BranchMasterRecord, ItemMasterRecord, MasterDataDiagnostics


def test_apply_approval_result_writes_governed_outputs_and_audit(tmp_path: Path) -> None:
    (tmp_path / "config").mkdir()
    result = ApprovalResult(
        approved_branches=(
            {
                "source_code": "HH",
                "branch_source_code": "01",
                "branch_source_name": "Test",
                "sap_card_code": "CHH-001",
                "mapping_status": "APPROVED",
                "approval_reference": "EMAIL-1",
            },
        ),
        approved_products=(
            {
                "source_code": "HH",
                "source_product_code": "ABC",
                "proposed_sap_item_code": "FA001",
                "recommended_action": "ADD_OSCN",
                "approval_reference": "EMAIL-1",
                "request_status": "READY_FOR_SAP_ADMIN",
            },
        ),
        rejected_branches=0,
        rejected_products=0,
        issues=(),
    )

    review = tmp_path / "review.xlsx"
    review.write_bytes(b"review evidence")
    assert apply_approval_result(tmp_path, result, review) == (1, 1)
    backups = sorted((tmp_path / "output" / "governance_backups").glob("*/manifest.json"))
    assert len(backups) == 1

    with (tmp_path / "config" / "branch_crosswalk.csv").open(
        encoding="utf-8-sig", newline=""
    ) as stream:
        assert list(csv.DictReader(stream))[0]["approval_reference"] == "EMAIL-1"
    with (tmp_path / "output" / "operations" / "oscn_change_requests.csv").open(
        encoding="utf-8-sig", newline=""
    ) as stream:
        assert list(csv.DictReader(stream))[0]["request_status"] == "READY_FOR_SAP_ADMIN"
    with (tmp_path / "output" / "operations" / "approval_audit_log.csv").open(
        encoding="utf-8-sig", newline=""
    ) as stream:
        audit = list(csv.DictReader(stream))
    assert {row["entity_type"] for row in audit} == {"BRANCH", "PRODUCT"}
    assert all(len(row["workbook_sha256"]) == 64 for row in audit)

    assert apply_approval_result(tmp_path, result, review) == (1, 1)
    with (tmp_path / "output" / "operations" / "approval_audit_log.csv").open(
        encoding="utf-8-sig", newline=""
    ) as stream:
        assert len(list(csv.DictReader(stream))) == 2


def test_apply_refuses_invalid_result(tmp_path: Path) -> None:
    from esip.approvals import ApprovalIssue

    result = ApprovalResult((), (), 0, 0, (ApprovalIssue("Sheet", 2, "bad"),))
    try:
        apply_approval_result(tmp_path, result)
    except ValueError:
        pass
    else:
        raise AssertionError("Expected invalid approval result to be rejected")


def test_workbook_fixture_has_expected_approval_columns(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Branch Approval"
    sheet.append(["mapping_status", "approval_reference"])
    sheet.append(["APPROVED", "EMAIL-1"])
    product = workbook.create_sheet("Product Mapping")
    product.append(["mapping_status", "approval_reference"])
    path = tmp_path / "review.xlsx"
    workbook.save(path)
    assert path.is_file()


def test_conflicting_duplicate_product_approvals_are_rejected(
    tmp_path: Path, monkeypatch
) -> None:
    (tmp_path / "config").mkdir()
    (tmp_path / "config" / "source_registry.yaml").write_text(
        "sources:\n  HH:\n    name: HomeHub\n", encoding="utf-8"
    )
    operations = tmp_path / "output" / "operations"
    operations.mkdir(parents=True)
    (operations / "product_mapping_queue.csv").write_text(
        "source_code,source_product_code\nHH,ABC\n", encoding="utf-8"
    )
    (operations / "branch_mapping_approval_queue.csv").write_text(
        "source_code,branch_source_code,branch_source_name\n", encoding="utf-8"
    )
    diagnostics = MasterDataDiagnostics(0, (), ())
    monkeypatch.setattr(
        approvals,
        "load_branch_master",
        lambda _path: ([BranchMasterRecord("CHH-001", "Test")], diagnostics),
    )
    monkeypatch.setattr(
        approvals,
        "load_item_master",
        lambda _path: (
            [
                ItemMasterRecord("FA001", "One", "", "Y", "ALUMINIUM"),
                ItemMasterRecord("FA002", "Two", "", "Y", "ALUMINIUM"),
            ],
            diagnostics,
        ),
    )
    workbook = Workbook()
    branch = workbook.active
    branch.title = "Branch Approval"
    branch.append(["mapping_status"])
    product = workbook.create_sheet("Product Mapping")
    product.append(
        [
            "source_code",
            "source_product_code",
            "candidate_sap_item_codes",
            "recommended_action",
            "mapping_status",
            "approval_reference",
        ]
    )
    product.append(["HH", "ABC", "FA001", "ADD_OSCN", "APPROVED", "EMAIL-1"])
    product.append(["HH", "ABC", "FA002", "ADD_OSCN", "APPROVED", "EMAIL-1"])
    path = tmp_path / "review.xlsx"
    workbook.save(path)

    result = approvals.evaluate_approval_workbook(tmp_path, path)

    assert any("conflicting targets" in issue.message for issue in result.issues)
