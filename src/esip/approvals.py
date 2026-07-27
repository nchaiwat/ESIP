from __future__ import annotations

import csv
from datetime import datetime, timezone
import hashlib
import os
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Iterable

from openpyxl import load_workbook

from esip.governance_backup import create_governance_backup
from esip.master_data import load_branch_master, load_item_master, normalize_identifier
from esip.profiles import load_yaml


BRANCH_FIELDS = (
    "source_code",
    "branch_source_code",
    "branch_source_name",
    "sap_card_code",
    "mapping_status",
    "approval_reference",
)
OSCN_REQUEST_FIELDS = (
    "source_code",
    "source_product_code",
    "proposed_sap_item_code",
    "recommended_action",
    "approval_reference",
    "request_status",
)
AUDIT_FIELDS = (
    "applied_at_utc",
    "workbook_path",
    "workbook_sha256",
    "entity_type",
    "source_code",
    "source_key",
    "target_sap_code",
    "approval_reference",
    "action",
)


@dataclass(frozen=True)
class ApprovalIssue:
    sheet: str
    row: int
    message: str


@dataclass(frozen=True)
class ApprovalResult:
    approved_branches: tuple[dict[str, str], ...]
    approved_products: tuple[dict[str, str], ...]
    rejected_branches: int
    rejected_products: int
    issues: tuple[ApprovalIssue, ...]


def _worksheet_rows(path: Path, sheet_name: str) -> Iterable[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing required worksheet: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_identifier(value) for value in next(rows)]
        for row_number, values in enumerate(rows, start=2):
            yield row_number, dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


def _normalized_status(value: Any) -> str:
    return normalize_identifier(value).upper() or "PENDING"


def _approval_text(value: Any) -> str:
    return normalize_identifier(value).removeprefix("'").replace("\u200b", "")


def _source_codes(root: Path) -> set[str]:
    registry = load_yaml(root / "config" / "source_registry.yaml")
    return {normalize_identifier(code).upper() for code in registry.get("sources", {})}


def _canonical_queue_keys(root: Path) -> tuple[set[tuple[str, str]], set[tuple[str, str, str]]]:
    product_rows = _read_csv(root / "output" / "operations" / "product_mapping_queue.csv")
    product_keys = {
        (
            normalize_identifier(row.get("source_code")).upper(),
            _approval_text(row.get("source_product_code")),
        )
        for row in product_rows
    }
    branch_rows = _read_csv(
        root / "output" / "operations" / "branch_mapping_approval_queue.csv"
    )
    branch_keys = {
        (
            normalize_identifier(row.get("source_code")).upper(),
            _approval_text(row.get("branch_source_code")),
            normalize_identifier(row.get("branch_source_name")),
        )
        for row in branch_rows
    }
    return product_keys, branch_keys


def evaluate_approval_workbook(root: Path, workbook_path: Path) -> ApprovalResult:
    issues: list[ApprovalIssue] = []
    approved_branches: list[dict[str, str]] = []
    approved_products: list[dict[str, str]] = []
    rejected_branches = 0
    rejected_products = 0
    source_codes = _source_codes(root)
    product_queue_keys, branch_queue_keys = _canonical_queue_keys(root)
    seen_branch_targets: dict[tuple[str, str, str], str] = {}
    seen_product_targets: dict[tuple[str, str], str] = {}

    branch_records, _ = load_branch_master(
        root / "MasterData" / "BranchMaster" / "incoming" / "ModernTrade_Branch.xlsx"
    )
    valid_card_codes = {record.card_code.upper() for record in branch_records}
    item_records, _ = load_item_master(
        root / "MasterData" / "ItemMaster" / "incoming" / "ItemMaster_FGFU_OITM.xlsx"
    )
    valid_item_codes = {record.item_code.upper() for record in item_records}

    try:
        branch_rows = list(_worksheet_rows(workbook_path, "Branch Approval"))
        product_rows = list(_worksheet_rows(workbook_path, "Product Mapping"))
    except (OSError, ValueError, StopIteration) as error:
        return ApprovalResult((), (), 0, 0, (ApprovalIssue("Workbook", 0, str(error)),))

    for row_number, row in branch_rows:
        status = _normalized_status(row.get("mapping_status"))
        if status == "PENDING":
            continue
        if status == "REJECTED":
            rejected_branches += 1
            continue
        if status != "APPROVED":
            issues.append(
                ApprovalIssue("Branch Approval", row_number, f"Unknown mapping_status: {status}")
            )
            continue
        source_code = normalize_identifier(row.get("source_code")).upper()
        branch_code = _approval_text(row.get("branch_source_code"))
        branch_name = normalize_identifier(row.get("branch_source_name"))
        card_code = normalize_identifier(row.get("candidate_card_code")).upper()
        reference = normalize_identifier(row.get("approval_reference"))
        row_issues: list[str] = []
        if source_code not in source_codes:
            row_issues.append(f"source_code {source_code or '<blank>'} is not registered")
        if not branch_code and not branch_name:
            row_issues.append("branch source code and name are both blank")
        if not card_code:
            row_issues.append("candidate_card_code is required")
        elif card_code not in valid_card_codes:
            row_issues.append(f"candidate_card_code {card_code} is not in Branch Master")
        if not reference:
            row_issues.append("approval_reference is required")
        branch_key = (source_code, branch_code, branch_name)
        if branch_key not in branch_queue_keys:
            row_issues.append("source branch identity is not in the current governed queue")
        prior_target = seen_branch_targets.get(branch_key)
        if prior_target and prior_target != card_code:
            row_issues.append(
                f"duplicate source branch has conflicting targets: {prior_target} and {card_code}"
            )
        if row_issues:
            issues.extend(
                ApprovalIssue("Branch Approval", row_number, message) for message in row_issues
            )
            continue
        seen_branch_targets[branch_key] = card_code
        approved_branches.append(
            {
                "source_code": source_code,
                "branch_source_code": branch_code,
                "branch_source_name": branch_name,
                "sap_card_code": card_code,
                "mapping_status": "APPROVED",
                "approval_reference": reference,
            }
        )

    for row_number, row in product_rows:
        status = _normalized_status(row.get("mapping_status"))
        if status == "PENDING":
            continue
        if status == "REJECTED":
            rejected_products += 1
            continue
        if status != "APPROVED":
            issues.append(
                ApprovalIssue("Product Mapping", row_number, f"Unknown mapping_status: {status}")
            )
            continue
        source_code = normalize_identifier(row.get("source_code")).upper()
        source_product_code = _approval_text(row.get("source_product_code"))
        candidates = [
            candidate.strip()
            for candidate in normalize_identifier(row.get("candidate_sap_item_codes")).split("|")
            if candidate.strip()
        ]
        reference = normalize_identifier(row.get("approval_reference"))
        row_issues = []
        if source_code not in source_codes:
            row_issues.append(f"source_code {source_code or '<blank>'} is not registered")
        if not source_product_code:
            row_issues.append("source_product_code is required")
        if len(candidates) != 1:
            row_issues.append("exactly one candidate_sap_item_code is required")
        elif candidates[0].upper() not in valid_item_codes:
            row_issues.append(f"candidate SAP ItemCode {candidates[0]} is not in Item Master")
        if not reference:
            row_issues.append("approval_reference is required")
        product_key = (source_code, source_product_code)
        if product_key not in product_queue_keys:
            row_issues.append("source product identity is not in the current governed queue")
        target = candidates[0] if len(candidates) == 1 else ""
        prior_target = seen_product_targets.get(product_key)
        if prior_target and prior_target != target:
            row_issues.append(
                f"duplicate source product has conflicting targets: {prior_target} and {target}"
            )
        if row_issues:
            issues.extend(
                ApprovalIssue("Product Mapping", row_number, message) for message in row_issues
            )
            continue
        seen_product_targets[product_key] = candidates[0]
        approved_products.append(
            {
                "source_code": source_code,
                "source_product_code": source_product_code,
                "proposed_sap_item_code": candidates[0],
                "recommended_action": normalize_identifier(row.get("recommended_action")),
                "approval_reference": reference,
                "request_status": "READY_FOR_SAP_ADMIN",
            }
        )

    return ApprovalResult(
        tuple(approved_branches),
        tuple(approved_products),
        rejected_branches,
        rejected_products,
        tuple(issues),
    )


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def _atomic_write_csv(
    path: Path, fieldnames: tuple[str, ...], rows: Iterable[dict[str, str]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(
        "w",
        encoding="utf-8-sig",
        newline="",
        dir=path.parent,
        delete=False,
        suffix=".tmp",
    ) as stream:
        temporary_path = Path(stream.name)
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        stream.flush()
        os.fsync(stream.fileno())
    temporary_path.replace(path)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def apply_approval_result(
    root: Path,
    result: ApprovalResult,
    workbook_path: Path | None = None,
) -> tuple[int, int]:
    if result.issues:
        raise ValueError("Approval result has validation issues; no files were changed")

    branch_path = root / "config" / "branch_crosswalk.csv"
    existing_branches = _read_csv(branch_path)
    branch_by_key = {
        (
            row.get("source_code", ""),
            row.get("branch_source_code", ""),
            row.get("branch_source_name", ""),
        ): row
        for row in existing_branches
        if row.get("mapping_status", "").upper() == "APPROVED"
        and row.get("approval_reference", "").strip()
    }
    for row in result.approved_branches:
        key = (row["source_code"], row["branch_source_code"], row["branch_source_name"])
        branch_by_key[key] = row

    product_path = root / "output" / "operations" / "oscn_change_requests.csv"
    existing_products = _read_csv(product_path)
    product_by_key = {
        (row.get("source_code", ""), row.get("source_product_code", "")): row
        for row in existing_products
    }
    for row in result.approved_products:
        product_by_key[(row["source_code"], row["source_product_code"])] = row

    workbook_hash = _sha256(workbook_path) if workbook_path else ""
    workbook_label = str(workbook_path.resolve()) if workbook_path else ""
    applied_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    audit_path = root / "output" / "operations" / "approval_audit_log.csv"
    audit_rows = _read_csv(audit_path)
    audit_keys = {
        (
            row.get("workbook_sha256", ""),
            row.get("entity_type", ""),
            row.get("source_code", ""),
            row.get("source_key", ""),
            row.get("target_sap_code", ""),
            row.get("approval_reference", ""),
        )
        for row in audit_rows
    }
    new_audit_rows: list[dict[str, str]] = []
    for entity_type, rows, source_field, target_field, action in (
        (
            "BRANCH",
            result.approved_branches,
            "branch_source_code",
            "sap_card_code",
            "UPSERT_BRANCH_CROSSWALK",
        ),
        (
            "PRODUCT",
            result.approved_products,
            "source_product_code",
            "proposed_sap_item_code",
            "CREATE_OSCN_CHANGE_REQUEST",
        ),
    ):
        for row in rows:
            source_key = row[source_field]
            key = (
                workbook_hash,
                entity_type,
                row["source_code"],
                source_key,
                row[target_field],
                row["approval_reference"],
            )
            if key in audit_keys:
                continue
            new_audit_rows.append(
                {
                    "applied_at_utc": applied_at,
                    "workbook_path": workbook_label,
                    "workbook_sha256": workbook_hash,
                    "entity_type": entity_type,
                    "source_code": row["source_code"],
                    "source_key": source_key,
                    "target_sap_code": row[target_field],
                    "approval_reference": row["approval_reference"],
                    "action": action,
                }
            )
            audit_keys.add(key)

    paths = (branch_path, product_path, audit_path)
    prior_contents = {path: path.read_bytes() if path.is_file() else None for path in paths}
    create_governance_backup(
        root,
        "MAPPING_APPROVAL",
        paths,
        workbook_path=workbook_path,
        data_snapshot={
            "approved_branch_count": len(result.approved_branches),
            "approved_product_count": len(result.approved_products),
        },
    )
    try:
        _atomic_write_csv(branch_path, BRANCH_FIELDS, branch_by_key.values())
        _atomic_write_csv(product_path, OSCN_REQUEST_FIELDS, product_by_key.values())
        _atomic_write_csv(audit_path, AUDIT_FIELDS, [*audit_rows, *new_audit_rows])
    except Exception:
        for path, content in prior_contents.items():
            if content is None:
                path.unlink(missing_ok=True)
            else:
                path.write_bytes(content)
        raise
    return len(result.approved_branches), len(result.approved_products)
