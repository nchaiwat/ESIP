from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import psycopg
from openpyxl import load_workbook

from esip.daily_ingest import (
    IngestSummary,
    _decimal,
    _oscn_index,
    existing_batch_summaries,
)
from esip.master_data import load_oscn
from esip.postgres import database_url


def gbh_branch_layout(
    branch_codes: tuple[object, ...],
    branch_names: tuple[object, ...],
    measure_headers: tuple[object, ...],
) -> list[tuple[str, str, int, int, int]]:
    """Return GBH branch columns as zero-based stock, quantity, and amount indexes."""
    starts = [
        index
        for index, value in enumerate(branch_codes)
        if str(value or "").strip() and str(value).strip() != "Grand Total"
    ]
    layout: list[tuple[str, str, int, int, int]] = []
    for position, start in enumerate(starts):
        stop = starts[position + 1] if position + 1 < len(starts) else len(measure_headers)
        columns = {
            str(measure_headers[index] or "").strip(): index
            for index in range(start, stop)
            if measure_headers[index]
        }
        required = ("สต็อค ณ ปัจจุบัน", "จำนวนขาย", "ยอดขายสุทธิ")
        if not all(name in columns for name in required):
            raise ValueError(f"Incomplete GBH measure block at column {start + 1}")
        layout.append(
            (
                str(branch_codes[start]).strip(),
                str(branch_names[start] or "").strip(),
                columns[required[0]],
                columns[required[1]],
                columns[required[2]],
            )
        )
    return layout


def report_date_from_filename(path: Path) -> date:
    match = re.search(r"(20\d{2})-(\d{2})-(\d{2})", path.name)
    if not match:
        raise ValueError(f"Report date not found in {path.name}")
    return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def _read_gbh_rows(
    path: Path,
    mapping: dict[str, str | None],
) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    next(iterator)
    branch_codes = next(iterator)
    branch_names = next(iterator)
    headers = next(iterator)
    layout = gbh_branch_layout(branch_codes, branch_names, headers)
    rows: list[dict[str, object]] = []
    for row_no, row in enumerate(iterator, start=5):
        product_code = str(row[0] or "").strip()
        if not product_code or product_code == "Grand Total":
            continue
        sap_item = mapping.get(product_code)
        for branch_code, branch_name, stock_col, qty_col, amount_col in layout:
            rows.append(
                {
                    "product_code": product_code,
                    "sap_item": sap_item,
                    "branch_code": branch_code,
                    "branch_name": branch_name,
                    "inventory_qty": _decimal(str(row[stock_col] or 0)),
                    "sales_qty": _decimal(str(row[qty_col] or 0)),
                    "sales_amount": _decimal(str(row[amount_col] or 0)),
                    "sheet": sheet.title,
                    "row_no": row_no,
                    "file_name": path.name,
                }
            )
    workbook.close()
    return rows


def _gbh_quarantine(
    row: dict[str, object],
    dataset: str,
    reason_code: str,
    reason_detail: str,
    batch_id: str,
) -> tuple[object, ...]:
    return (
        "GBH",
        dataset,
        reason_code,
        reason_detail,
        json.dumps(
            {
                "product_code": row["product_code"],
                "branch_code": row["branch_code"],
                "branch_name": row["branch_name"],
            },
            ensure_ascii=False,
        ),
        batch_id,
        row["file_name"],
        row["sheet"],
        row["row_no"],
    )


def _persist_gbh_date(
    root: Path,
    business_date: date,
    paths: list[Path],
    rows: list[dict[str, object]],
) -> list[IngestSummary]:
    batch_id, digest, source_file_name = _gbh_batch_identity(business_date, paths)
    with psycopg.connect(database_url(root)) as connection:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1 FROM import_batch WHERE import_batch_id=%s", (batch_id,))
            if cursor.fetchone():
                return [
                    IngestSummary(
                        "GBH",
                        dataset,
                        len(rows),
                        0,
                        0,
                        Decimal(),
                        Decimal(),
                        Decimal(),
                        batch_id,
                        True,
                    )
                    for dataset in ("sales", "inventory")
                ]
            cursor.execute(
                "INSERT INTO import_batch VALUES (%s,'GBH',%s,%s,%s,'RECEIVED')",
                (
                    batch_id,
                    source_file_name,
                    digest,
                    datetime.now(timezone.utc),
                ),
            )
            cursor.execute(
                """INSERT INTO batch_governance VALUES
                (%s,'DAILY_RAW','EVIDENCE_BASED_DRAFT','PENDING_ADMIN',NULL)""",
                (batch_id,),
            )
            summaries: list[IngestSummary] = []
            for dataset in ("sales", "inventory"):
                measure_key = "sales_amount" if dataset == "sales" else "inventory_qty"
                source_measure = sum(
                    (row[measure_key] for row in rows),
                    Decimal(),
                )
                facts: list[tuple[object, ...]] = []
                quarantine: list[tuple[object, ...]] = []
                quarantine_measure = Decimal()
                seen: set[tuple[object, object]] = set()
                for row in rows:
                    measure = row[measure_key]
                    grain = (row["product_code"], row["branch_code"])
                    if grain in seen:
                        quarantine.append(
                            _gbh_quarantine(
                                row,
                                dataset,
                                "DUPLICATE_SOURCE_GRAIN",
                                "Product and branch repeated across GBH partitions",
                                batch_id,
                            )
                        )
                        quarantine_measure += measure
                        continue
                    seen.add(grain)
                    if not row["sap_item"]:
                        quarantine.append(
                            _gbh_quarantine(
                                row,
                                dataset,
                                "PRODUCT_NOT_MAPPED",
                                "No unique CGH OSCN match for source product code",
                                batch_id,
                            )
                        )
                        quarantine_measure += measure
                        continue
                    common = (
                        "GBH",
                        business_date,
                        row["branch_code"],
                        row["branch_name"],
                        row["product_code"],
                        row["sap_item"],
                    )
                    lineage = (
                        batch_id,
                        row["file_name"],
                        row["sheet"],
                        row["row_no"],
                    )
                    if dataset == "sales":
                        kind = (
                            "RETURN"
                            if row["sales_qty"] < 0 or row["sales_amount"] < 0
                            else "SALE"
                        )
                        facts.append(
                            common
                            + (row["sales_qty"], row["sales_amount"], kind)
                            + lineage
                        )
                    else:
                        facts.append(
                            common + (row["inventory_qty"], None) + lineage
                        )
                if dataset == "sales":
                    cursor.executemany(
                        """INSERT INTO fact_sales (
                        source_code,sales_date,branch_source_code,branch_source_name,
                        product_source_code,sap_item_code,sales_qty,
                        sales_amount_ex_vat_after_discount,record_type,import_batch_id,
                        source_file_name,source_sheet_name,source_row_no)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        facts,
                    )
                else:
                    cursor.executemany(
                        """INSERT INTO fact_inventory_snapshot (
                        source_code,snapshot_date,branch_source_code,branch_source_name,
                        product_source_code,sap_item_code,onhand_qty,onhand_value,
                        import_batch_id,source_file_name,source_sheet_name,source_row_no)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        facts,
                    )
                cursor.executemany(
                    """INSERT INTO quarantine_record (
                    source_code,dataset,reason_code,reason_detail,source_payload_json,
                    import_batch_id,source_file_name,source_sheet_name,source_row_no)
                    VALUES (%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s)""",
                    quarantine,
                )
                staged_measure = source_measure - quarantine_measure
                passed = source_measure == staged_measure + quarantine_measure
                cursor.execute(
                    """INSERT INTO batch_reconciliation (
                    import_batch_id,dataset,source_rows,staged_rows,quarantined_rows,
                    source_measure,staged_measure,quarantined_measure,passed)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        batch_id,
                        dataset,
                        len(rows),
                        len(facts),
                        len(quarantine),
                        source_measure,
                        staged_measure,
                        quarantine_measure,
                        passed,
                    ),
                )
                summaries.append(
                    IngestSummary(
                        "GBH",
                        dataset,
                        len(rows),
                        len(facts),
                        len(quarantine),
                        source_measure,
                        staged_measure,
                        quarantine_measure,
                        batch_id,
                    )
                )
            cursor.execute(
                """UPDATE import_batch SET status = CASE WHEN
                (SELECT BOOL_AND(passed) FROM batch_reconciliation
                 WHERE import_batch_id=%s) THEN 'RECONCILED' ELSE 'QUARANTINED' END
                WHERE import_batch_id=%s""",
                (batch_id, batch_id),
            )
    return summaries


def _gbh_batch_identity(
    business_date: date, paths: list[Path]
) -> tuple[str, str, str]:
    identity = "|".join(
        f"{path.name}:{hashlib.sha256(path.read_bytes()).hexdigest().upper()}"
        for path in sorted(paths)
    )
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest().upper()
    return (
        f"GBH-combined-{business_date.isoformat()}-{digest[:12]}",
        digest,
        "|".join(path.name for path in sorted(paths)),
    )


def ingest_gbh(root: Path) -> list[IngestSummary]:
    incoming = root / "SourceFiles" / "GBH" / "incoming"
    grouped: dict[date, list[Path]] = {}
    for path in sorted(incoming.glob("*.xlsx")):
        grouped.setdefault(report_date_from_filename(path), []).append(path)
    mapping: dict[str, str | None] | None = None
    summaries: list[IngestSummary] = []
    for business_date, paths in sorted(grouped.items()):
        batch_id, _, _ = _gbh_batch_identity(business_date, paths)
        existing = existing_batch_summaries(root, batch_id)
        if existing:
            summaries.extend(existing)
            continue
        if mapping is None:
            mapping = _oscn(root, "CGH")
        rows = [
            row
            for path in paths
            for row in _read_gbh_rows(path, mapping)
        ]
        summaries.extend(_persist_gbh_date(root, business_date, paths, rows))
    return summaries


def _oscn(root: Path, prefix: str) -> dict[str, str | None]:
    path = (
        root / "MasterData" / "OSCN" / "incoming"
        / "All OSCN Before add 6 Row for all CTW req by Pun 2026-06-19.xlsx"
    )
    records, _ = load_oscn(path)
    return _oscn_index(records, prefix)


def _date_from_text(value: object, fallback: date) -> date:
    match = re.search(r"20\d{2}-\d{2}-\d{2}", str(value))
    return date.fromisoformat(match.group()) if match else fallback


def _existing_for_path(
    root: Path, source_code: str, dataset: str, path: Path
) -> list[IngestSummary]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest().upper()
    return existing_batch_summaries(root, f"{source_code}-{dataset}-{digest[:16]}")


def _persist(
    root: Path,
    source_code: str,
    dataset: str,
    path: Path,
    business_date: date,
    rows: list[dict[str, object]],
    reconciliation_measure_key: str | None = None,
) -> IngestSummary:
    digest = hashlib.sha256(path.read_bytes()).hexdigest().upper()
    batch_id = f"{source_code}-{dataset}-{digest[:16]}"
    measure_key = reconciliation_measure_key or (
        "amount" if dataset == "sales" else "qty"
    )
    source_measure = sum((row[measure_key] for row in rows), Decimal())
    facts, quarantine = [], []
    quarantine_measure = Decimal()
    for row in rows:
        measure = row[measure_key]
        if not row["sap_item"]:
            quarantine.append(
                (
                    source_code, dataset, "PRODUCT_NOT_MAPPED",
                    "No unique OSCN match for source product code",
                    json.dumps(
                        {
                            "product_code": row["product_code"],
                            "branch_code": row["branch_code"],
                        },
                        ensure_ascii=False,
                    ),
                    batch_id, path.name, row["sheet"], row["row_no"],
                )
            )
            quarantine_measure += measure
            continue
        common = (
            source_code, business_date, row["branch_code"], row["branch_name"],
            row["product_code"], row["sap_item"],
        )
        if dataset == "sales":
            kind = "RETURN" if row["qty"] < 0 or row["amount"] < 0 else "SALE"
            facts.append(
                common + (
                    row["qty"], row["amount"], kind, batch_id, path.name,
                    row["sheet"], row["row_no"],
                )
            )
        else:
            facts.append(
                common + (
                    row["qty"], row.get("value"), batch_id, path.name,
                    row["sheet"], row["row_no"],
                )
            )
    staged_measure = source_measure - quarantine_measure
    with psycopg.connect(database_url(root)) as connection:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1 FROM import_batch WHERE import_batch_id=%s", (batch_id,))
            if cursor.fetchone():
                return IngestSummary(
                    source_code, dataset, len(rows), 0, 0, source_measure,
                    Decimal(), Decimal(), batch_id, True,
                )
            cursor.execute(
                "INSERT INTO import_batch VALUES (%s,%s,%s,%s,%s,'RECEIVED')",
                (batch_id, source_code, path.name, digest, datetime.now(timezone.utc)),
            )
            cursor.execute(
                """INSERT INTO batch_governance VALUES
                (%s,'DAILY_RAW','EVIDENCE_BASED_DRAFT','PENDING_ADMIN',NULL)""",
                (batch_id,),
            )
            if dataset == "sales":
                cursor.executemany(
                    """INSERT INTO fact_sales (
                    source_code,sales_date,branch_source_code,branch_source_name,
                    product_source_code,sap_item_code,sales_qty,
                    sales_amount_ex_vat_after_discount,record_type,import_batch_id,
                    source_file_name,source_sheet_name,source_row_no)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    facts,
                )
            else:
                cursor.executemany(
                    """INSERT INTO fact_inventory_snapshot (
                    source_code,snapshot_date,branch_source_code,branch_source_name,
                    product_source_code,sap_item_code,onhand_qty,onhand_value,
                    import_batch_id,source_file_name,source_sheet_name,source_row_no)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    facts,
                )
            cursor.executemany(
                """INSERT INTO quarantine_record (
                source_code,dataset,reason_code,reason_detail,source_payload_json,
                import_batch_id,source_file_name,source_sheet_name,source_row_no)
                VALUES (%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s)""",
                quarantine,
            )
            passed = source_measure == staged_measure + quarantine_measure
            cursor.execute(
                """INSERT INTO batch_reconciliation (
                import_batch_id,dataset,source_rows,staged_rows,quarantined_rows,
                source_measure,staged_measure,quarantined_measure,passed)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    batch_id, dataset, len(rows), len(facts), len(quarantine),
                    source_measure, staged_measure, quarantine_measure, passed,
                ),
            )
            cursor.execute(
                "UPDATE import_batch SET status=%s WHERE import_batch_id=%s",
                ("RECONCILED" if passed else "QUARANTINED", batch_id),
            )
    return IngestSummary(
        source_code, dataset, len(rows), len(facts), len(quarantine),
        source_measure, staged_measure, quarantine_measure, batch_id,
    )


def ingest_hh(root: Path) -> list[IngestSummary]:
    incoming = root / "SourceFiles" / "HH" / "incoming"
    mapping: dict[str, str | None] | None = None
    summaries = []
    patterns = (("sales", "SaleReport*.xlsx"), ("inventory", "StockReport*.xlsx"))
    for dataset, pattern in patterns:
        for path in sorted(incoming.rglob(pattern)):
            existing = _existing_for_path(root, "HH", dataset, path)
            if existing:
                summaries.extend(existing)
                continue
            if mapping is None:
                mapping = _oscn(root, "CHH")
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            values = list(sheet.iter_rows(values_only=True))
            fallback = date.fromtimestamp(path.stat().st_mtime)
            report_date = _date_from_text(values[0][0], fallback)
            branches = [
                (column, str(values[1][column] or "").strip())
                for column in range(6, 16, 2)
                if values[1][column]
            ]
            rows = []
            for row_no, row in enumerate(values[3:], start=4):
                product_code = str(row[3] or "").strip()
                if not product_code:
                    continue
                sap_item = mapping.get(product_code)
                for column, branch_name in branches:
                    qty = _decimal(str(row[column] or 0))
                    second = _decimal(str(row[column + 1] or 0))
                    rows.append(
                        {
                            "product_code": product_code,
                            "sap_item": sap_item,
                            "branch_code": branch_name,
                            "branch_name": branch_name,
                            "qty": qty,
                            "amount": second if dataset == "sales" else Decimal(),
                            "value": second if dataset == "inventory" else None,
                            "sheet": sheet.title,
                            "row_no": row_no,
                        }
                    )
            workbook.close()
            summaries.append(_persist(root, "HH", dataset, path, report_date, rows))
    return summaries


def _ingest_dh_inventory_file(root: Path, path: Path) -> IngestSummary:
    existing = _existing_for_path(root, "DH", "inventory", path)
    if existing:
        return existing[0]
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    mapping = _oscn(root, "CDH")
    rows = []
    for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        product_code = str(row[2] or "").strip()
        if not product_code:
            continue
        rows.append(
            {
                "product_code": product_code,
                "sap_item": mapping.get(product_code),
                "branch_code": str(row[0] or "").strip(),
                "branch_name": str(row[1] or "").strip(),
                "qty": _decimal(str(row[8] or 0)),
                "amount": Decimal(),
                "value": None,
                "sheet": sheet.title,
                "row_no": row_no,
            }
        )
    workbook.close()
    match = re.search(r"(\d{2})-(\d{2})-(20\d{2})", path.name)
    if not match:
        raise RuntimeError(f"Snapshot date not found in {path.name}")
    snapshot_date = date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    return _persist(root, "DH", "inventory", path, snapshot_date, rows)


def ingest_dh_inventory(root: Path) -> list[IngestSummary]:
    incoming = root / "SourceFiles" / "DH" / "incoming"
    return [
        _ingest_dh_inventory_file(root, path)
        for path in sorted(incoming.glob("*สต็อค*.xlsx"))
    ]


def _ingest_dh_sales_file(root: Path, path: Path) -> IngestSummary:
    existing = _existing_for_path(root, "DH", "sales", path)
    if existing:
        return existing[0]
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator)
    branches = [
        (index, str(value).strip())
        for index, value in enumerate(headers[3:], start=3)
        if value and str(value).strip() != "รวม"
    ]
    mapping = _oscn(root, "CDH")
    rows: list[dict[str, object]] = []
    for row_no, row in enumerate(iterator, start=2):
        product_code = str(row[0] or "").strip()
        if not product_code:
            continue
        sap_item = mapping.get(product_code)
        for column, branch_name in branches:
            rows.append(
                {
                    "product_code": product_code,
                    "sap_item": sap_item,
                    "branch_code": branch_name,
                    "branch_name": branch_name,
                    "qty": _decimal(str(row[column] or 0)),
                    "amount": Decimal(),
                    "value": None,
                    "sheet": sheet.title,
                    "row_no": row_no,
                }
            )
    workbook.close()
    match = re.search(r"(\d{2})-(\d{2})-(20\d{2})", path.name)
    if not match:
        raise RuntimeError(f"Sales date not found in {path.name}")
    sales_date = date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    return _persist(
        root,
        "DH",
        "sales",
        path,
        sales_date,
        rows,
        reconciliation_measure_key="qty",
    )


def ingest_dh_sales(root: Path) -> list[IngestSummary]:
    incoming = root / "SourceFiles" / "DH" / "incoming"
    return [
        _ingest_dh_sales_file(root, path)
        for path in sorted(incoming.glob("*ยอดขาย*.xlsx"))
    ]
