from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from esip.manifest import sha256_file
from esip.master_data import OscnRecord, normalize_identifier
from esip.profiles import load_yaml


@dataclass(frozen=True)
class StagingSummary:
    source_code: str
    source_rows: int
    filtered_rows: int
    staged_rows: int
    quarantined_rows: int
    source_qty: Decimal
    staged_qty: Decimal
    quarantined_qty: Decimal
    source_amount: Decimal
    staged_amount: Decimal
    quarantined_amount: Decimal

    @property
    def reconciles(self) -> bool:
        return (
            self.filtered_rows == self.staged_rows + self.quarantined_rows
            and self.source_qty == self.staged_qty + self.quarantined_qty
            and self.source_amount == self.staged_amount + self.quarantined_amount
        )


def parse_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    try:
        return Decimal(str(value))
    except InvalidOperation as error:
        raise ValueError(f"invalid decimal {value!r}") from error


def parse_sales_date(values: dict[str, Any]) -> date:
    direct = values.get("sales_date")
    if isinstance(direct, datetime):
        return direct.date()
    if isinstance(direct, date):
        return direct
    if direct:
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(str(direct).strip(), pattern).date()
            except ValueError:
                pass
    if values.get("sales_year") and values.get("sales_month"):
        return date(
            int(values["sales_year"]),
            int(values["sales_month"]),
            int(values.get("sales_day") or 1),
        )
    period = str(values.get("sales_period") or "").strip()
    for pattern in ("%b'%y", "%m/%Y", "%d/%m/%Y", "%Y-%m"):
        try:
            return datetime.strptime(period, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"invalid sales date/period {direct or period!r}")


def build_oscn_index(
    records: list[OscnRecord], cardcode_prefix: str
) -> dict[str, tuple[str, ...]]:
    mapped: dict[str, set[str]] = defaultdict(set)
    for record in records:
        if record.card_code.startswith(cardcode_prefix):
            mapped[record.customer_sku].add(record.item_code)
    return {key: tuple(sorted(values)) for key, values in mapped.items()}


def _include_row(source_code: str, branch_code: str) -> bool:
    if source_code == "HP":
        return not branch_code.startswith("M")
    if source_code == "MH":
        return branch_code.startswith("M")
    return True


def stage_sales(
    workspace: Path,
    source_code: str,
    workbook_path: Path,
    oscn_records: list[OscnRecord],
    output_dir: Path,
) -> StagingSummary:
    profile = load_yaml(workspace / "ImportProfiles" / f"{source_code}.yaml")
    registry = load_yaml(workspace / "config" / "source_registry.yaml")
    dataset = next(iter(profile["datasets"].values()))
    columns = dataset["column_positions"]
    oscn_index = build_oscn_index(
        oscn_records, registry["sources"][source_code]["sap_cardcode_prefix"]
    )
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[dataset["sheet"]]
    batch_id = f"{source_code}-{sha256_file(workbook_path)[:16]}"
    output_dir.mkdir(parents=True, exist_ok=True)
    staged_path = output_dir / f"{source_code.lower()}_canonical_sales.csv"
    quarantine_path = output_dir / f"{source_code.lower()}_quarantine.csv"
    staged_fields = [
        "source_code",
        "sales_date",
        "branch_source_code",
        "branch_source_name",
        "product_source_code",
        "product_source_name",
        "sap_item_code",
        "sales_qty",
        "sales_amount_ex_vat_after_discount",
        "record_type",
        "import_batch_id",
        "source_file_name",
        "source_sheet_name",
        "source_row_no",
        "mapping_status",
    ]
    quarantine_fields = [
        "source_code",
        "reason_code",
        "reason_detail",
        "import_batch_id",
        "source_file_name",
        "source_sheet_name",
        "source_row_no",
        "source_payload_json",
    ]
    source_rows = filtered_rows = staged_rows = quarantined_rows = 0
    source_qty = staged_qty = quarantined_qty = Decimal("0")
    source_amount = staged_amount = quarantined_amount = Decimal("0")
    with (
        staged_path.open("w", encoding="utf-8-sig", newline="") as staged_stream,
        quarantine_path.open("w", encoding="utf-8-sig", newline="") as quarantine_stream,
    ):
        staged_writer = csv.DictWriter(staged_stream, fieldnames=staged_fields)
        quarantine_writer = csv.DictWriter(quarantine_stream, fieldnames=quarantine_fields)
        staged_writer.writeheader()
        quarantine_writer.writeheader()
        rows = sheet.iter_rows(min_row=dataset["data_start_row"], values_only=True)
        for row_no, row in enumerate(rows, start=dataset["data_start_row"]):
            source_rows += 1
            values = {
                field: row[position - 1] if position <= len(row) else None
                for field, position in columns.items()
            }
            branch_code = normalize_identifier(values.get("branch_source_code"))
            if not _include_row(source_code, branch_code):
                continue
            filtered_rows += 1
            try:
                product_code = normalize_identifier(values.get("product_source_code"))
                if not product_code:
                    raise ValueError("product_source_code is blank")
                sales_date = parse_sales_date(values)
                qty = parse_decimal(values.get("sales_qty"))
                amount = parse_decimal(values.get("sales_amount_ex_vat_after_discount"))
                source_qty += qty
                source_amount += amount
                matches = oscn_index.get(product_code, ())
                if not matches:
                    reason_code = "UNMAPPED_PRODUCT"
                    reason_detail = "No SAP OSCN match for partner prefix and customer SKU"
                elif len(matches) > 1:
                    reason_code = "AMBIGUOUS_PRODUCT"
                    reason_detail = f"Customer SKU maps to SAP ItemCodes: {'|'.join(matches)}"
                else:
                    staged_writer.writerow(
                        {
                            "source_code": source_code,
                            "sales_date": sales_date.isoformat(),
                            "branch_source_code": branch_code,
                            "branch_source_name": normalize_identifier(values.get("branch_source_name")),
                            "product_source_code": product_code,
                            "product_source_name": normalize_identifier(values.get("product_source_name")),
                            "sap_item_code": matches[0],
                            "sales_qty": str(qty),
                            "sales_amount_ex_vat_after_discount": str(amount),
                            "record_type": "RETURN" if qty < 0 or amount < 0 else "SALE",
                            "import_batch_id": batch_id,
                            "source_file_name": workbook_path.name,
                            "source_sheet_name": sheet.title,
                            "source_row_no": row_no,
                            "mapping_status": "MAPPED_OSCN_PREFIX",
                        }
                    )
                    staged_rows += 1
                    staged_qty += qty
                    staged_amount += amount
                    continue
            except (TypeError, ValueError) as error:
                reason_code = "INVALID_ROW"
                reason_detail = str(error)
                qty = amount = Decimal("0")
            quarantine_writer.writerow(
                {
                    "source_code": source_code,
                    "reason_code": reason_code,
                    "reason_detail": reason_detail,
                    "import_batch_id": batch_id,
                    "source_file_name": workbook_path.name,
                    "source_sheet_name": sheet.title,
                    "source_row_no": row_no,
                    "source_payload_json": json.dumps(values, ensure_ascii=False, default=str),
                }
            )
            quarantined_rows += 1
            quarantined_qty += qty
            quarantined_amount += amount
    workbook.close()
    return StagingSummary(
        source_code=source_code,
        source_rows=source_rows,
        filtered_rows=filtered_rows,
        staged_rows=staged_rows,
        quarantined_rows=quarantined_rows,
        source_qty=source_qty,
        staged_qty=staged_qty,
        quarantined_qty=quarantined_qty,
        source_amount=source_amount,
        staged_amount=staged_amount,
        quarantined_amount=quarantined_amount,
    )
