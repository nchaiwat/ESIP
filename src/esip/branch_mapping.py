from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

from esip.master_data import BranchMasterRecord, normalize_identifier
from esip.profiles import load_yaml
from esip.staging import _include_row


@dataclass(frozen=True)
class SourceBranch:
    source_code: str
    branch_source_code: str
    branch_source_name: str


@dataclass(frozen=True)
class BranchCandidate:
    source: SourceBranch
    card_code: str
    card_name: str
    score: float
    rank: int
    recommendation: str


def normalize_branch_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u0080-\uffff]+", "", value).casefold()


def branch_similarity(source_name: str, card_name: str) -> float:
    source = normalize_branch_name(source_name)
    target = normalize_branch_name(card_name)
    if not source or not target:
        return 0.0
    score = SequenceMatcher(None, source, target).ratio()
    if len(source) >= 3 and source in target:
        score = max(score, 0.95)
    return round(score, 4)


def collect_source_branches(
    workspace: Path, source_code: str, workbook_path: Path
) -> list[SourceBranch]:
    profile = load_yaml(workspace / "ImportProfiles" / f"{source_code}.yaml")
    dataset = next(iter(profile["datasets"].values()))
    columns = dataset["column_positions"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[dataset["sheet"]]
    branches: set[tuple[str, str]] = set()
    for row in sheet.iter_rows(min_row=dataset["data_start_row"], values_only=True):
        code_position = columns.get("branch_source_code")
        name_position = columns.get("branch_source_name") or columns.get("branch_source_name_raw")
        code = normalize_identifier(row[code_position - 1]) if code_position else ""
        name = normalize_identifier(row[name_position - 1]) if name_position else ""
        if _include_row(source_code, code) and (code or name):
            branches.add((code, name))
    workbook.close()
    return [SourceBranch(source_code, code, name) for code, name in sorted(branches)]


def rank_branch_candidates(
    source: SourceBranch,
    master_records: list[BranchMasterRecord],
    cardcode_prefix: str,
    limit: int = 3,
) -> list[BranchCandidate]:
    candidates = [record for record in master_records if record.card_code.startswith(cardcode_prefix)]
    ranked = sorted(
        ((branch_similarity(source.branch_source_name, record.card_name), record) for record in candidates),
        key=lambda item: (-item[0], item[1].card_code),
    )[:limit]
    if not ranked:
        return []
    margin = ranked[0][0] - ranked[1][0] if len(ranked) > 1 else ranked[0][0]
    output: list[BranchCandidate] = []
    for rank, (score, record) in enumerate(ranked, start=1):
        recommendation = "REVIEW"
        if rank == 1 and score >= 0.92 and margin >= 0.05:
            recommendation = "HIGH_CONFIDENCE_CANDIDATE"
        output.append(
            BranchCandidate(source, record.card_code, record.card_name, score, rank, recommendation)
        )
    return output


def write_branch_candidate_report(candidates: list[BranchCandidate], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(
            [
                "source_code",
                "branch_source_code",
                "branch_source_name",
                "candidate_card_code",
                "candidate_card_name",
                "score",
                "rank",
                "recommendation",
            ]
        )
        for candidate in candidates:
            writer.writerow(
                [
                    candidate.source.source_code,
                    candidate.source.branch_source_code,
                    candidate.source.branch_source_name,
                    candidate.card_code,
                    candidate.card_name,
                    candidate.score,
                    candidate.rank,
                    candidate.recommendation,
                ]
            )
