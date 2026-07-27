from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg

from esip.governance_backup import create_governance_backup
from esip.postgres import database_url
from esip.profiles import load_yaml


QUEUE_FIELDS = (
    "import_batch_id",
    "source_code",
    "source_file_name",
    "imported_at_utc",
    "batch_status",
    "profile_evidence_status",
    "reconciliation_passed",
    "quarantine_rows",
    "source_branch_identities",
    "approved_branch_mappings",
    "branch_mapping_rate",
    "readiness_status",
    "blocking_reasons",
    "approval_status",
    "approval_reference",
)
AUDIT_FIELDS = (
    "applied_at_utc",
    "workbook_sha256",
    "import_batch_id",
    "source_code",
    "approval_reference",
    "action",
)


@dataclass(frozen=True)
class PublicationIssue:
    row: int
    message: str


@dataclass(frozen=True)
class PublicationApprovalResult:
    approved: tuple[tuple[str, str, str], ...]
    rejected: int
    issues: tuple[PublicationIssue, ...]


def publication_rows(root: Path) -> list[dict[str, Any]]:
    profiles = {
        path.stem: load_yaml(path).get("status", "")
        for path in (root / "ImportProfiles").glob("*.yaml")
    }
    with psycopg.connect(database_url(root)) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """WITH branch_identity AS (
                    SELECT import_batch_id, source_code, branch_source_code,
                           COALESCE(branch_source_name,'') AS branch_source_name
                    FROM fact_sales
                    UNION
                    SELECT import_batch_id, source_code, branch_source_code,
                           COALESCE(branch_source_name,'')
                    FROM fact_inventory_snapshot
                ), branch_coverage AS (
                    SELECT i.import_batch_id,
                           COUNT(*) AS source_branch_identities,
                           COUNT(b.sap_card_code) AS approved_branch_mappings
                    FROM branch_identity i
                    LEFT JOIN bridge_source_branch b
                      ON b.source_code=i.source_code
                     AND b.branch_source_code=i.branch_source_code
                     AND b.branch_source_name=i.branch_source_name
                     AND b.mapping_status='APPROVED'
                    GROUP BY i.import_batch_id
                ), recon AS (
                    SELECT import_batch_id, BOOL_AND(passed) AS passed
                    FROM batch_reconciliation GROUP BY import_batch_id
                ), quarantine AS (
                    SELECT import_batch_id, COUNT(*) AS rows
                    FROM quarantine_record GROUP BY import_batch_id
                )
                SELECT b.import_batch_id,b.source_code,b.source_file_name,b.imported_at_utc,
                       b.status,COALESCE(r.passed,FALSE),COALESCE(q.rows,0),
                       COALESCE(c.source_branch_identities,0),
                       COALESCE(c.approved_branch_mappings,0)
                FROM import_batch b
                LEFT JOIN recon r USING(import_batch_id)
                LEFT JOIN quarantine q USING(import_batch_id)
                LEFT JOIN branch_coverage c USING(import_batch_id)
                ORDER BY b.imported_at_utc,b.source_code,b.import_batch_id"""
            )
            records = cursor.fetchall()
    rows: list[dict[str, Any]] = []
    for (
        batch_id,
        source_code,
        source_file,
        imported_at,
        batch_status,
        reconciled,
        quarantine_rows,
        branch_total,
        branch_approved,
    ) in records:
        profile_status = profiles.get(source_code, "missing")
        reasons: list[str] = []
        if batch_status != "RECONCILED":
            reasons.append("BATCH_NOT_RECONCILED")
        if not reconciled:
            reasons.append("RECONCILIATION_NOT_PASSED")
        if profile_status != "validated_daily_raw":
            reasons.append("DAILY_RAW_PROFILE_NOT_VALIDATED")
        if quarantine_rows:
            reasons.append("QUARANTINE_NOT_ZERO")
        if branch_total == 0 or branch_approved != branch_total:
            reasons.append("BRANCH_MAPPING_INCOMPLETE")
        rate = branch_approved / branch_total if branch_total else 0
        rows.append(
            {
                "import_batch_id": batch_id,
                "source_code": source_code,
                "source_file_name": source_file,
                "imported_at_utc": imported_at.isoformat(),
                "batch_status": batch_status,
                "profile_evidence_status": profile_status,
                "reconciliation_passed": bool(reconciled),
                "quarantine_rows": quarantine_rows,
                "source_branch_identities": branch_total,
                "approved_branch_mappings": branch_approved,
                "branch_mapping_rate": rate,
                "readiness_status": "READY_FOR_APPROVAL" if not reasons else "BLOCKED",
                "blocking_reasons": "|".join(reasons),
                "approval_status": "PENDING",
                "approval_reference": "",
            }
        )
    return rows


def export_publication_queue(root: Path) -> tuple[int, int]:
    rows = publication_rows(root)
    path = root / "output" / "operations" / "publication_readiness_queue.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=QUEUE_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows), sum(row["readiness_status"] == "READY_FOR_APPROVAL" for row in rows)


def _sheet_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Publication Readiness" not in workbook.sheetnames:
            raise ValueError("Missing required worksheet: Publication Readiness")
        sheet = workbook["Publication Readiness"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [
            (row_number, dict(zip(headers, row, strict=False)))
            for row_number, row in enumerate(values, start=2)
        ]
    finally:
        workbook.close()


def evaluate_publication_approvals(
    root: Path, workbook_path: Path
) -> PublicationApprovalResult:
    current = {row["import_batch_id"]: row for row in publication_rows(root)}
    approved: list[tuple[str, str, str]] = []
    rejected = 0
    issues: list[PublicationIssue] = []
    try:
        workbook_rows = _sheet_rows(workbook_path)
    except (OSError, ValueError, StopIteration) as error:
        return PublicationApprovalResult((), 0, (PublicationIssue(0, str(error)),))
    seen: set[str] = set()
    for row_number, row in workbook_rows:
        status = str(row.get("approval_status") or "PENDING").strip().upper()
        if status == "PENDING":
            continue
        if status == "REJECTED":
            rejected += 1
            continue
        if status != "APPROVED":
            issues.append(PublicationIssue(row_number, f"Unknown approval_status: {status}"))
            continue
        batch_id = str(row.get("import_batch_id") or "").strip()
        reference = str(row.get("approval_reference") or "").strip()
        live = current.get(batch_id)
        if batch_id in seen:
            issues.append(PublicationIssue(row_number, "Duplicate approved batch"))
        elif not live:
            issues.append(PublicationIssue(row_number, "Batch is not in the current queue"))
        elif live["readiness_status"] != "READY_FOR_APPROVAL":
            issues.append(
                PublicationIssue(
                    row_number,
                    f"Batch is blocked: {live['blocking_reasons']}",
                )
            )
        elif not reference:
            issues.append(PublicationIssue(row_number, "approval_reference is required"))
        else:
            approved.append((batch_id, live["source_code"], reference))
            seen.add(batch_id)
    return PublicationApprovalResult(tuple(approved), rejected, tuple(issues))


def _workbook_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def apply_publication_approvals(
    root: Path,
    workbook_path: Path,
    result: PublicationApprovalResult,
) -> int:
    if result.issues:
        raise ValueError("Publication approval validation failed")
    if not result.approved:
        return 0
    live = {row["import_batch_id"]: row for row in publication_rows(root)}
    for batch_id, _, _ in result.approved:
        if live.get(batch_id, {}).get("readiness_status") != "READY_FOR_APPROVAL":
            raise ValueError(f"Publication readiness changed before apply: {batch_id}")
    applied_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    workbook_hash = _workbook_sha256(workbook_path)
    audit_path = root / "output" / "operations" / "publication_approval_audit.csv"
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    existing: list[dict[str, str]] = []
    if audit_path.is_file():
        with audit_path.open(encoding="utf-8-sig", newline="") as stream:
            existing = list(csv.DictReader(stream))
    keys = {
        (row["workbook_sha256"], row["import_batch_id"], row["approval_reference"])
        for row in existing
    }
    new_rows = []
    with psycopg.connect(database_url(root)) as connection:
        with connection.cursor() as cursor:
            batch_ids = [batch_id for batch_id, _, _ in result.approved]
            cursor.execute(
                """SELECT import_batch_id, input_classification, profile_status,
                branch_mapping_status, approval_reference
                FROM batch_governance
                WHERE import_batch_id = ANY(%s)
                ORDER BY import_batch_id""",
                (batch_ids,),
            )
            columns = [column.name for column in cursor.description or ()]
            before_rows = [
                dict(zip(columns, row, strict=True)) for row in cursor.fetchall()
            ]
            create_governance_backup(
                root,
                "PUBLICATION_APPROVAL",
                [audit_path],
                workbook_path=workbook_path,
                data_snapshot={"batch_governance_before": before_rows},
            )
            for batch_id, source_code, reference in result.approved:
                cursor.execute(
                    """UPDATE batch_governance
                    SET profile_status='APPROVED',
                        branch_mapping_status='APPROVED',
                        approval_reference=%s
                    WHERE import_batch_id=%s""",
                    (reference, batch_id),
                )
                if cursor.rowcount != 1:
                    raise RuntimeError(f"Batch governance row not found: {batch_id}")
                key = (workbook_hash, batch_id, reference)
                if key not in keys:
                    new_rows.append(
                        {
                            "applied_at_utc": applied_at,
                            "workbook_sha256": workbook_hash,
                            "import_batch_id": batch_id,
                            "source_code": source_code,
                            "approval_reference": reference,
                            "action": "APPROVE_BATCH_GOVERNANCE_NOT_PUBLISH",
                        }
                    )
                    keys.add(key)
    with audit_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=AUDIT_FIELDS)
        writer.writeheader()
        writer.writerows([*existing, *new_rows])
    return len(result.approved)
