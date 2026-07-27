from __future__ import annotations

from collections import Counter
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass(frozen=True)
class ItemMasterRecord:
    item_code: str
    item_name: str
    barcode: str
    active: str
    product_family: str


def classify_product_family(item_code: str) -> str:
    normalized = item_code.strip().upper()
    if normalized.startswith("FA"):
        return "ALUMINIUM"
    if normalized.startswith("FU"):
        return "UPVC"
    return "OUT_OF_SCOPE"


@dataclass(frozen=True)
class OscnRecord:
    item_code: str
    card_code: str
    customer_sku: str
    partner_barcode: str


@dataclass(frozen=True)
class BranchMasterRecord:
    card_code: str
    card_name: str


@dataclass(frozen=True)
class MasterDataDiagnostics:
    row_count: int
    blank_key_rows: tuple[int, ...]
    duplicate_keys: tuple[tuple[str, ...], ...]
    ambiguous_keys: tuple[tuple[str, ...], ...] = ()


def _read_rows(path: Path, sheet_name: str | None = None) -> Iterable[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_identifier(value) for value in next(rows)]
    for row_no, values in enumerate(rows, start=2):
        yield row_no, dict(zip(headers, values, strict=False))
    workbook.close()


def _diagnostics(keys: list[tuple[str, ...]], blank_rows: list[int]) -> MasterDataDiagnostics:
    counts = Counter(key for key in keys if all(key))
    duplicates = tuple(sorted(key for key, count in counts.items() if count > 1))
    return MasterDataDiagnostics(len(keys), tuple(blank_rows), duplicates)


def load_item_master(path: Path) -> tuple[list[ItemMasterRecord], MasterDataDiagnostics]:
    records: list[ItemMasterRecord] = []
    keys: list[tuple[str, ...]] = []
    blank_rows: list[int] = []
    for row_no, row in _read_rows(path):
        item_code = normalize_identifier(row.get("Item No."))
        keys.append((item_code,))
        if not item_code:
            blank_rows.append(row_no)
            continue
        records.append(
            ItemMasterRecord(
                item_code=item_code,
                item_name=normalize_identifier(row.get("Item Description")),
                barcode=normalize_identifier(row.get("Bar Code")),
                active=normalize_identifier(row.get("Active")),
                product_family=classify_product_family(item_code),
            )
        )
    return records, _diagnostics(keys, blank_rows)


def load_oscn(path: Path) -> tuple[list[OscnRecord], MasterDataDiagnostics]:
    records: list[OscnRecord] = []
    keys: list[tuple[str, ...]] = []
    blank_rows: list[int] = []
    item_codes_by_key: dict[tuple[str, str], set[str]] = {}
    for row_no, row in _read_rows(path, "AllOSCN"):
        item_code = normalize_identifier(row.get("Item No."))
        card_code = normalize_identifier(row.get("BP Code"))
        customer_sku = normalize_identifier(row.get("BP Catalog Number"))
        key = (card_code, customer_sku)
        keys.append(key)
        if not all(key) or not item_code:
            blank_rows.append(row_no)
            continue
        item_codes_by_key.setdefault(key, set()).add(item_code)
        records.append(
            OscnRecord(
                item_code=item_code,
                card_code=card_code,
                customer_sku=customer_sku,
                partner_barcode=normalize_identifier(row.get("Partner Barcode")),
            )
        )
    base = _diagnostics(keys, blank_rows)
    ambiguous = tuple(sorted(key for key, values in item_codes_by_key.items() if len(values) > 1))
    return records, MasterDataDiagnostics(
        row_count=base.row_count,
        blank_key_rows=base.blank_key_rows,
        duplicate_keys=base.duplicate_keys,
        ambiguous_keys=ambiguous,
    )


def load_branch_master(path: Path) -> tuple[list[BranchMasterRecord], MasterDataDiagnostics]:
    records: list[BranchMasterRecord] = []
    keys: list[tuple[str, ...]] = []
    blank_rows: list[int] = []
    for row_no, row in _read_rows(path):
        card_code = normalize_identifier(row.get("BP Code"))
        keys.append((card_code,))
        if not card_code:
            blank_rows.append(row_no)
            continue
        records.append(
            BranchMasterRecord(
                card_code=card_code,
                card_name=normalize_identifier(row.get("BP Name")),
            )
        )
    return records, _diagnostics(keys, blank_rows)


def write_oscn_ambiguity_report(records: list[OscnRecord], path: Path) -> int:
    grouped: dict[tuple[str, str], set[str]] = {}
    for record in records:
        grouped.setdefault((record.card_code, record.customer_sku), set()).add(record.item_code)
    ambiguous = [(key, sorted(values)) for key, values in grouped.items() if len(values) > 1]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(["card_code", "customer_sku", "sap_item_codes", "item_code_count"])
        for (card_code, customer_sku), item_codes in sorted(ambiguous):
            writer.writerow([card_code, customer_sku, "|".join(item_codes), len(item_codes)])
    return len(ambiguous)
