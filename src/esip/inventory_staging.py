from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from esip.manifest import sha256_file
from esip.master_data import OscnRecord, normalize_identifier
from esip.profiles import load_yaml
from esip.staging import build_oscn_index, parse_decimal, parse_sales_date


@dataclass(frozen=True)
class InventorySummary:
    source_code: str
    source_rows: int
    staged_rows: int
    quarantined_rows: int
    duplicate_grain_rows: int
    source_onhand_qty: Decimal
    staged_onhand_qty: Decimal
    quarantined_onhand_qty: Decimal

    @property
    def reconciles(self) -> bool:
        return (
            self.source_rows == self.staged_rows + self.quarantined_rows
            and self.source_onhand_qty
            == self.staged_onhand_qty + self.quarantined_onhand_qty
        )


@dataclass(frozen=True)
class InventoryInputRow:
    row_no: int
    snapshot_date: date
    branch_code: str
    branch_name: str
    product_code: str
    product_name: str
    onhand_qty: Decimal
    payload: dict[str, Any]


def parse_snapshot_from_sheet(sheet_name: str) -> date:
    value = sheet_name.removeprefix("Stock ").strip()
    return datetime.strptime(value, "%d.%m.%y").date()


def stage_inventory(
    workspace: Path,
    source_code: str,
    workbook_path: Path,
    oscn_records: list[OscnRecord],
    output_dir: Path,
) -> InventorySummary:
    profile = load_yaml(workspace / "ImportProfiles" / f"{source_code}.yaml")
    registry = load_yaml(workspace / "config" / "source_registry.yaml")
    dataset_name = "inventory" if "inventory" in profile["datasets"] else "sales_inventory"
    dataset = profile["datasets"][dataset_name]
    columns = dataset["column_positions"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "sheet" in dataset:
        sheet = workbook[dataset["sheet"]]
    else:
        pattern_prefix = dataset["sheet_pattern"].removesuffix("*")
        sheet = workbook[next(name for name in workbook.sheetnames if name.startswith(pattern_prefix))]
    oscn_index = build_oscn_index(
        oscn_records, registry["sources"][source_code]["sap_cardcode_prefix"]
    )
    batch_id = f"{source_code}-{sha256_file(workbook_path)[:16]}"
    rows_by_grain: dict[tuple[date, str, str], list[InventoryInputRow]] = defaultdict(list)
    source_onhand = Decimal("0")
    for row_no, row in enumerate(
        sheet.iter_rows(min_row=dataset["data_start_row"], values_only=True),
        start=dataset["data_start_row"],
    ):
        values = {
            field: row[position - 1] if position <= len(row) else None
            for field, position in columns.items()
        }
        try:
            snapshot_date = (
                parse_snapshot_from_sheet(sheet.title)
                if dataset.get("snapshot_date_from") == "sheet_name"
                else parse_sales_date(values)
            )
            onhand = parse_decimal(values.get("onhand_qty"))
        except (TypeError, ValueError):
            snapshot_date = date.min
            onhand = Decimal("0")
        input_row = InventoryInputRow(
            row_no=row_no,
            snapshot_date=snapshot_date,
            branch_code=normalize_identifier(values.get("branch_source_code")),
            branch_name=normalize_identifier(
                values.get("branch_source_name") or values.get("branch_source_name_raw")
            ),
            product_code=normalize_identifier(values.get("product_source_code")),
            product_name=normalize_identifier(values.get("product_source_name")),
            onhand_qty=onhand,
            payload=values,
        )
        rows_by_grain[(snapshot_date, input_row.branch_code, input_row.product_code)].append(
            input_row
        )
        source_onhand += onhand

    output_dir.mkdir(parents=True, exist_ok=True)
    staged_path = output_dir / f"{source_code.lower()}_canonical_inventory.csv"
    quarantine_path = output_dir / f"{source_code.lower()}_inventory_quarantine.csv"
    staged_fields = [
        "source_code",
        "snapshot_date",
        "branch_source_code",
        "branch_source_name",
        "product_source_code",
        "product_source_name",
        "sap_item_code",
        "onhand_qty",
        "import_batch_id",
        "source_file_name",
        "source_sheet_name",
        "source_row_no",
        "mapping_status",
    ]
    quarantine_fields = [
        "source_code",
        "reason_code",
        "reason_detail",
        "import_batch_id",
        "source_file_name",
        "source_sheet_name",
        "source_row_no",
        "source_payload_json",
    ]
    staged_rows = quarantined_rows = duplicate_rows = 0
    staged_onhand = quarantined_onhand = Decimal("0")
    with (
        staged_path.open("w", encoding="utf-8-sig", newline="") as staged_stream,
        quarantine_path.open("w", encoding="utf-8-sig", newline="") as quarantine_stream,
    ):
        staged_writer = csv.DictWriter(staged_stream, fieldnames=staged_fields)
        quarantine_writer = csv.DictWriter(quarantine_stream, fieldnames=quarantine_fields)
        staged_writer.writeheader()
        quarantine_writer.writeheader()
        for grain, grouped_rows in rows_by_grain.items():
            duplicate = len(grouped_rows) > 1
            for input_row in grouped_rows:
                matches = oscn_index.get(input_row.product_code, ())
                if input_row.snapshot_date == date.min or not input_row.product_code:
                    reason_code = "INVALID_ROW"
                    reason_detail = "Invalid snapshot date or blank product code"
                elif duplicate:
                    reason_code = "DUPLICATE_INVENTORY_GRAIN"
                    reason_detail = (
                        f"{len(grouped_rows)} rows share snapshot/branch/product grain {grain}"
                    )
                    duplicate_rows += 1
                elif not matches:
                    reason_code = "UNMAPPED_PRODUCT"
                    reason_detail = "No SAP OSCN match for partner prefix and customer SKU"
                elif len(matches) > 1:
                    reason_code = "AMBIGUOUS_PRODUCT"
                    reason_detail = f"Customer SKU maps to SAP ItemCodes: {'|'.join(matches)}"
                else:
                    staged_writer.writerow(
                        {
                            "source_code": source_code,
                            "snapshot_date": input_row.snapshot_date.isoformat(),
                            "branch_source_code": input_row.branch_code,
                            "branch_source_name": input_row.branch_name,
                            "product_source_code": input_row.product_code,
                            "product_source_name": input_row.product_name,
                            "sap_item_code": matches[0],
                            "onhand_qty": str(input_row.onhand_qty),
                            "import_batch_id": batch_id,
                            "source_file_name": workbook_path.name,
                            "source_sheet_name": sheet.title,
                            "source_row_no": input_row.row_no,
                            "mapping_status": "MAPPED_OSCN_PREFIX",
                        }
                    )
                    staged_rows += 1
                    staged_onhand += input_row.onhand_qty
                    continue
                quarantine_writer.writerow(
                    {
                        "source_code": source_code,
                        "reason_code": reason_code,
                        "reason_detail": reason_detail,
                        "import_batch_id": batch_id,
                        "source_file_name": workbook_path.name,
                        "source_sheet_name": sheet.title,
                        "source_row_no": input_row.row_no,
                        "source_payload_json": json.dumps(
                            input_row.payload, ensure_ascii=False, default=str
                        ),
                    }
                )
                quarantined_rows += 1
                quarantined_onhand += input_row.onhand_qty
    workbook.close()
    return InventorySummary(
        source_code=source_code,
        source_rows=sum(len(rows) for rows in rows_by_grain.values()),
        staged_rows=staged_rows,
        quarantined_rows=quarantined_rows,
        duplicate_grain_rows=duplicate_rows,
        source_onhand_qty=source_onhand,
        staged_onhand_qty=staged_onhand,
        quarantined_onhand_qty=quarantined_onhand,
    )
